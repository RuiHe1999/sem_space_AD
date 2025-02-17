# 1. packages
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

import pingouin as pg
import statsmodels.api as sm
from scipy import stats
from statsmodels.stats import multitest

import seaborn as sns
sns.set(font='Arial')

from matplotlib import pyplot as plt
plt.rcParams["font.family"] = "Arial"

import warnings
warnings.filterwarnings("ignore")

# 3. functions
def add_sig(r, p):
    
    if p < 0.001:
        p = f'{r:.3f}***'
    elif p < 0.01:
        p = f'{r:.3f}**'
    elif p < 0.05:
        p = f'{r:.3f}*' 
         
    else:
        p = f'{r:.3f}'
    # p = f'{r:.3f}' 
    return p

def compute_corr(df, xs, ys, method, padjust, covar):
    
    df = df.copy()
    
    coefs = []
    annots = []
    for x in xs:
        rs = []
        qs = []
        rqs = []
        if x in ['FT_WordNum', 'FT_TTR', 'BERT_TokenNum', 'BERT_TTR']:
            ys_ = [y for y in ys if x!=y]
            corrs = pg.pairwise_corr(df, [[x], ys_], method=method, padjust=padjust, covar=covar)
        else:
            corrs = pg.pairwise_corr(df, [[x], ys], method=method, padjust=padjust, covar=covar)
            
        for y in ys:
            
            if len(corrs[corrs['Y'] == y]) != 0:
                r = corrs[corrs['Y'] == y].r.item()
                if padjust != None:
                    q = corrs[corrs['Y'] == y]['p-corr'].item()
                else:
                    q = corrs[corrs['Y'] == y]['p-unc'].item()
                    
                rs.append(r)
                qs.append(q)
                rqs.append(add_sig(r, q))
                
            else:
                rs.append(0)
                qs.append(1)
                rqs.append(' / ')
            
                
        rs = np.array(rs)    
        qs = np.array(qs)
        rqs = np.array(rqs)
        
        
        # threshold
        rs[qs >= 0.05] = 0
        coefs.append(rs)
        annots.append(rqs)
        
    coefs = np.array(coefs)
    annots = np.array(annots)
    
    return coefs, annots

def compute_glm(feat, formula, df):
    
    family=sm.families.Gaussian(link=sm.families.links.Identity())
    if feat=='PPL':
        family=sm.families.Gamma(link=sm.families.links.Log())
        
    # fit the model 
    model = sm.GLM.from_formula(formula, data=df, family=family)
    result = model.fit()

    # Deviance goodness-of-fit test
    deviance = result.deviance
    df_resid = result.df_resid
    fit_p = 1 - stats.chi2.cdf(deviance, df_resid)
    assert (fit_p > 0.05)

    # store all results in a dataframe
    summary = pd.DataFrame({"Var": result.params.index, 
                            "Coefficient": result.params, 
                            "CI_lower": result.conf_int()[0], 
                            "CI_upper": result.conf_int()[1],
                            "SE": result.bse, 
                            "z": result.tvalues,
                            "P>|z|": result.pvalues, 
                            })
    
    summary.index = range(summary.shape[0])
    
    summary.insert(0, 'Feat', [feat] * summary.shape[0])
    summary.insert(1, 'Deviance', [deviance] * summary.shape[0])
    summary.insert(2, 'Fit_p', [fit_p] * summary.shape[0])
 
    return result, summary

def set_font_for_entire_workbook():
    workbook = load_workbook('results.xlsx')
    times_new_roman_font = Font(name='Calibri')
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = times_new_roman_font
    
    workbook.save('results.xlsx')

# 3. commands    
# read results 
data = pd.read_csv('data.csv')
data.gender = data.gender.apply(lambda x: x.lower())
data['diagnosis'] = data.dx.map({'NC': 0, 'pAD': 1})
data['lang'] = data.language.map({'EL': 0, 'EN': 1})
data['sex'] = data.gender.map({'male': 0, 'female': 1})

en_data = data[data['language']=='EN']
el_data = data[data['language']=='EL']

variables = ['FT_GSim', 'FT_LSim', 'BERT_GSim', 'BERT_LSim', 'PPL', 'CLIP', 'ADD']
desp = ['age', 'educ', 'sex', 'FT_WordNum', 'FT_TTR', 'BERT_TokenNum', 'BERT_TTR']

# 3.1 correlations
coefs, annots = compute_corr(df=data, 
                             xs=desp, ys=['FT_WordNum', 'FT_TTR', 'BERT_TokenNum', 'BERT_TTR'] + variables, 
                             covar=['diagnosis', 'lang'],
                             method='spearman', padjust='fdr_bh',)

fig, ax = plt.subplots(1, 1, figsize=(12, 9))

sns.heatmap(coefs, fmt="", 
            annot=annots,  annot_kws={"size": 13}, 
            cbar=True, cbar_kws={"shrink": 0.8, "orientation": 'vertical'}, 
            vmin =-1, vmax = 1, 
            square=True, linewidth=.5, 
            cmap='coolwarm',
            ax=ax
            )
ax.set_xlabel("", fontsize=12)
ax.set_ylabel("", fontsize=12)
ax.set_xticklabels(['FT_WordNum', 'FT_TTR', 'BERT_TokenNum', 'BERT_TTR'] + variables, rotation=45, fontsize=12)
ax.set_yticklabels(desp, rotation=0, fontsize=12)
ax.xaxis.tick_top()

fig.savefig('correlation.png', dpi=600, bbox_inches='tight')

# 3.2 group comparions
# 3.2.1 GLM
formulas = {
    'FT_GSim':      'FT_GSim    ~  C(dx, Treatment("NC")) + C(sex) + FT_TTR',
    'FT_LSim':      'FT_LSim    ~  C(dx, Treatment("NC")) + FT_TTR',
    'BERT_GSim':    'BERT_GSim  ~  C(dx, Treatment("NC")) + BERT_TokenNum',
    'BERT_LSim':    'BERT_LSim  ~  C(dx, Treatment("NC"))',
    'CLIP':         'CLIP       ~  C(dx, Treatment("NC")) + BERT_TokenNum',
    'PPL':          'PPL        ~  C(dx, Treatment("NC")) + BERT_TTR',
    'ADD':          'ADD        ~  C(dx, Treatment("NC")) + FT_WordNum'
    }

en_summary = pd.DataFrame()
for variable in variables:
    formula = formulas[variable]
    feat_result, feat_summary = compute_glm(variable, formula, en_data)
    en_summary = pd.concat([en_summary, feat_summary])

el_summary = pd.DataFrame()
for variable in variables:
    formula = formulas[variable]
    feat_result, feat_summary = compute_glm(variable, formula, el_data)
    el_summary = pd.concat([el_summary, feat_summary])

en_summary.to_excel('results.xlsx', sheet_name='en_summary', index=False)
with pd.ExcelWriter('results.xlsx', engine='openpyxl', mode='a') as writer:
    el_summary.to_excel(writer, sheet_name='el_summary', index=False)

# 3.1.2 FDR correction for group comparison
en_glm_result = en_summary[en_summary['Var']=='C(dx, Treatment("NC"))[T.pAD]']
en_glm_result = en_glm_result.drop(columns=['Deviance', 'Fit_p', 'Var'])
en_glm_result['q'] = multitest.fdrcorrection(en_glm_result['P>|z|'])[1]
en_glm_result['langauge'] = 'EN'
    
el_glm_result = el_summary[el_summary['Var']=='C(dx, Treatment("NC"))[T.pAD]']
el_glm_result = el_glm_result.drop(columns=['Deviance', 'Fit_p', 'Var'])
el_glm_result['q'] = multitest.fdrcorrection(el_glm_result['P>|z|'])[1]
el_glm_result['langauge'] = 'EL'

group_results = pd.concat([en_glm_result, el_glm_result])

with pd.ExcelWriter('results.xlsx', engine='openpyxl', mode='a') as writer:
    group_results.to_excel(writer, sheet_name='group_comparison', index=False)


# 3.2.2 visualization 
fig, axes = plt.subplots(4, 4, figsize=(26, 26))
axes = axes.flatten()

for i, feat in enumerate(variables):
    # boxplot
    sns.boxplot(
        data=data, x="language", y=feat, hue="dx",
        palette=["#40A578", "#E65C19"], 
        fill=False, gap=.1,
        ax = axes[i*2]
    )

    axes[i*2].set_xlabel("", fontsize=14)
    axes[i*2].set_ylabel(feat, fontsize=14)
    axes[i*2].legend(loc=1, fontsize=14)
    axes[i*2].set_xticklabels(['English', 'Greek'], fontsize=14)
    axes[i*2].set_yticklabels(axes[i*2].get_yticklabels(), fontsize=14)
    axes[i*2].set_title(f'({chr(65+i*2)}) Group difference in {feat}', fontsize=16)

    # error plot 
    en_ci = group_results[(group_results['Feat']==feat) & (group_results['langauge']=='EN')]
    el_ci = group_results[(group_results['Feat']==feat) & (group_results['langauge']=='EL')]
    cis = {'EN': (en_ci.Coefficient.item(), en_ci.CI_lower.item(), en_ci.CI_upper.item()),
           'EL': (el_ci.Coefficient.item(), el_ci.CI_lower.item(), el_ci.CI_upper.item()),}
    cis = pd.DataFrame(cis).T
    cis.columns = ['coef', 'lower', 'upper']
    cis.insert(0, 'language', cis.index)

    axes[i*2+1].errorbar(cis['language'], cis['coef'], 
                        yerr=[cis['coef'] - cis['lower'], cis['upper'] - cis['coef']], 
                        fmt='o', capsize=5, ms=3, elinewidth=2, markeredgewidth=2,
                        ecolor='#e18727ff', mfc='#7e6148ff', mec='#7e6148ff')

    axes[i*2+1].set_xlim(-0.5, 1.5)
    axes[i*2+1].set_title(f'({chr(65+i*2+1)}) Error bar for coefficients for {feat}', fontsize=14)
    axes[i*2+1].axhline(en_ci.Coefficient.item(), color='gray', linestyle='dashed', linewidth=2)

for ax in axes[-2:]:
    ax.axis('off')

fig.savefig('group_comparison.jpg', dpi=600, bbox_inches='tight')

# 3.3. Syntax correlation 
# 3.3.1 GLM
formulas = {
    'FT_GSim':      'FT_GSim    ~  ADD + C(dx, Treatment("NC"))',
    'FT_LSim':      'FT_LSim    ~  ADD + C(dx, Treatment("NC"))',
    'BERT_GSim':    'BERT_GSim  ~  ADD + C(dx, Treatment("NC"))',
    'BERT_LSim':    'BERT_LSim  ~  ADD + C(dx, Treatment("NC"))',
    'CLIP':         'CLIP       ~  ADD + C(dx, Treatment("NC"))',
    'PPL':          'PPL        ~  ADD + C(dx, Treatment("NC"))',
    }


en_summary = pd.DataFrame()
for variable in variables[:-1]:
    formula = formulas[variable]
    feat_result, feat_summary = compute_glm(variable, formula, en_data)
    en_summary = pd.concat([en_summary, feat_summary])

el_summary = pd.DataFrame()
for variable in variables[:-1]:
    formula = formulas[variable]
    feat_result, feat_summary = compute_glm(variable, formula, el_data)
    el_summary = pd.concat([el_summary, feat_summary])

with pd.ExcelWriter('results.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    en_summary.to_excel(writer, sheet_name='en_sem_syn', index=False)
with pd.ExcelWriter('results.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    el_summary.to_excel(writer, sheet_name='el_sem_syn', index=False)

# 3.3.2 FDR correction for group comparison
en_glm_result = en_summary[en_summary['Var']=='ADD']
en_glm_result = en_glm_result.drop(columns=['Deviance', 'Fit_p', 'Var'])
en_glm_result['q'] = multitest.fdrcorrection(en_glm_result['P>|z|'])[1]
en_glm_result['langauge'] = 'EN'
    
el_glm_result = el_summary[el_summary['Var']=='ADD']
el_glm_result = el_glm_result.drop(columns=['Deviance', 'Fit_p', 'Var'])
el_glm_result['q'] = multitest.fdrcorrection(el_glm_result['P>|z|'])[1]
el_glm_result['langauge'] = 'EL'

add_results = pd.concat([en_glm_result, el_glm_result])

with pd.ExcelWriter('results.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    add_results.to_excel(writer, sheet_name='sem_syn', index=False) 

# 3.4.3. visualization
fig, axes = plt.subplots(1, 2, figsize=(20, 10))
axes = axes.flatten()
sig_feats = ['BERT_LSim', 'CLIP']

for i, feat in enumerate(sig_feats):
    
    # regression plot
    sns.regplot(
        data=en_data, x="ADD", y=feat,
        scatter_kws={'color': "#97BE5A", 's': 20},
        line_kws={'color': "#97BE5A"}, label='EN', ax=axes[i]
    )
    
    sns.regplot(
        data=el_data, x="ADD", y=feat,
        scatter_kws={'color': "#FFA27F", 's': 20},
        line_kws={'color': "#FFA27F"}, label='EL', ax=axes[i]
    )
    
    axes[i].set_xlabel("ADD", fontsize=18)
    axes[i].set_ylabel(feat, fontsize=18)
    axes[i].legend(loc=0, fontsize=18)
    axes[i].set_xticklabels(axes[i].get_xticklabels(), fontsize=18)
    axes[i].set_yticklabels(axes[i].get_yticklabels(), fontsize=18)
    axes[i].set_title(f'({chr(65+i)}) Relation between {feat} and ADD', fontsize=20)

fig.savefig('sem_syn_corr.jpg', dpi=600, bbox_inches='tight')    

# 3.3. Syntax correlation 
# 3.3.1 GLM
formulas = {
    'FT_GSim':      'FT_GSim    ~  ADD + C(dx, Treatment("NC"))',
    'FT_LSim':      'FT_LSim    ~  ADD + C(dx, Treatment("NC"))',
    'BERT_GSim':    'BERT_GSim  ~  ADD + C(dx, Treatment("NC"))',
    'BERT_LSim':    'BERT_LSim  ~  ADD + C(dx, Treatment("NC"))',
    'CLIP':         'CLIP       ~  ADD + C(dx, Treatment("NC"))',
    'PPL':          'PPL        ~  ADD + C(dx, Treatment("NC"))',
    }


en_summary = pd.DataFrame()
for variable in variables[:-1]:
    formula = formulas[variable]
    feat_result, feat_summary = compute_glm(variable, formula, en_data)
    en_summary = pd.concat([en_summary, feat_summary])

el_summary = pd.DataFrame()
for variable in variables[:-1]:
    formula = formulas[variable]
    feat_result, feat_summary = compute_glm(variable, formula, el_data)
    el_summary = pd.concat([el_summary, feat_summary])

with pd.ExcelWriter('results.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    en_summary.to_excel(writer, sheet_name='en_sem_syn', index=False)
with pd.ExcelWriter('results.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    el_summary.to_excel(writer, sheet_name='el_sem_syn', index=False)

# 3.3.2 FDR correction for group comparison
en_glm_result = en_summary[en_summary['Var']=='ADD']
en_glm_result = en_glm_result.drop(columns=['Deviance', 'Fit_p', 'Var'])
en_glm_result['q'] = multitest.fdrcorrection(en_glm_result['P>|z|'])[1]
en_glm_result['langauge'] = 'EN'
    
el_glm_result = el_summary[el_summary['Var']=='ADD']
el_glm_result = el_glm_result.drop(columns=['Deviance', 'Fit_p', 'Var'])
el_glm_result['q'] = multitest.fdrcorrection(el_glm_result['P>|z|'])[1]
el_glm_result['langauge'] = 'EL'

add_results = pd.concat([en_glm_result, el_glm_result])

with pd.ExcelWriter('results.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    add_results.to_excel(writer, sheet_name='sem_syn', index=False) 

# 3.4.3. visualization
fig, axes = plt.subplots(1, 2, figsize=(20, 10))
axes = axes.flatten()
sig_feats = ['BERT_LSim', 'CLIP']

for i, feat in enumerate(sig_feats):
    
    # regression plot
    sns.regplot(
        data=en_data, x="ADD", y=feat,
        scatter_kws={'color': "#97BE5A", 's': 20},
        line_kws={'color': "#97BE5A"}, label='EN', ax=axes[i]
    )
    
    sns.regplot(
        data=el_data, x="ADD", y=feat,
        scatter_kws={'color': "#FFA27F", 's': 20},
        line_kws={'color': "#FFA27F"}, label='EL', ax=axes[i]
    )
    
    axes[i].set_xlabel("ADD", fontsize=18)
    axes[i].set_ylabel(feat, fontsize=18)
    axes[i].legend(loc=0, fontsize=18)
    axes[i].set_xticklabels(axes[i].get_xticklabels(), fontsize=18)
    axes[i].set_yticklabels(axes[i].get_yticklabels(), fontsize=18)
    axes[i].set_title(f'({chr(65+i)}) Relation between {feat} and ADD', fontsize=20)

fig.savefig('sem_syn_corr.jpg', dpi=600, bbox_inches='tight') 

# 3.4. MMSE correlation
en_data = en_data[en_data['dx']=='pAD']
el_data = el_data[el_data['dx']=='pAD']

# 3.4.1 GLM
formulas = {
    'FT_GSim':      'FT_GSim    ~  mmse + C(sex) + FT_TTR',
    'FT_LSim':      'FT_LSim    ~  mmse + FT_TTR',
    'BERT_GSim':    'BERT_GSim  ~  mmse + BERT_TokenNum',
    'BERT_LSim':    'BERT_LSim  ~  mmse',
    'CLIP':         'CLIP       ~  mmse + BERT_TokenNum',
    'PPL':          'PPL        ~  mmse + BERT_TTR',
    'ADD':          'ADD        ~  mmse + FT_WordNum'
    }

en_summary = pd.DataFrame()
for variable in variables:
    formula = formulas[variable]
    feat_result, feat_summary = compute_glm(variable, formula, en_data)
    en_summary = pd.concat([en_summary, feat_summary])

el_summary = pd.DataFrame()
for variable in variables:
    formula = formulas[variable]
    feat_result, feat_summary = compute_glm(variable, formula, el_data)
    el_summary = pd.concat([el_summary, feat_summary])

with pd.ExcelWriter('results.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    en_summary.to_excel(writer, sheet_name='en_mmse', index=False)
with pd.ExcelWriter('results.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    el_summary.to_excel(writer, sheet_name='el_mmse', index=False)

# 3.4.2 FDR correction for group comparison
en_glm_result = en_summary[en_summary['Var']=='mmse']
en_glm_result = en_glm_result.drop(columns=['Deviance', 'Fit_p', 'Var'])
en_glm_result['q'] = multitest.fdrcorrection(en_glm_result['P>|z|'])[1]
en_glm_result['langauge'] = 'EN'
    
el_glm_result = el_summary[el_summary['Var']=='mmse']
el_glm_result = el_glm_result.drop(columns=['Deviance', 'Fit_p', 'Var'])
el_glm_result['q'] = multitest.fdrcorrection(el_glm_result['P>|z|'])[1]
el_glm_result['langauge'] = 'EL'

mmse_results = pd.concat([en_glm_result, el_glm_result])

with pd.ExcelWriter('results.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    mmse_results.to_excel(writer, sheet_name='mmse', index=False) 
    
set_font_for_entire_workbook()

# 3.4.3. visualization
fig, axes = plt.subplots(1, 2, figsize=(20, 10))
axes = axes.flatten()
sig_feats = ['FT_GSim']

for i, feat in enumerate(sig_feats):
    
    # regression plot
    sns.regplot(
        data=en_data, x="mmse", y=feat,
        scatter_kws={'color': "#97BE5A", 's': 20},
        line_kws={'color': "#97BE5A"}, label='EN', ax=axes[i*2]
    )

    sns.regplot(
        data=el_data, x="mmse", y=feat,
        scatter_kws={'color': "#FFA27F", 's': 20},
        line_kws={'color': "#FFA27F"}, label='EL', ax=axes[i*2]
    )


    axes[i*2].set_xlabel("MMSE", fontsize=18)
    axes[i*2].set_ylabel(feat, fontsize=18)
    axes[i*2].legend(loc=0, fontsize=18)
    axes[i*2].set_xticklabels(axes[i*2].get_xticklabels(), fontsize=18)
    axes[i*2].set_yticklabels(axes[i*2].get_yticklabels(), fontsize=18)
    axes[i*2].set_title(f'({chr(65+i*2)}) Relation between {feat} and MMSE score', fontsize=20)

    # error plot 
    en_ci = mmse_results[(mmse_results['Feat']==feat) & (mmse_results['langauge']=='EN')]
    el_ci = mmse_results[(mmse_results['Feat']==feat) & (mmse_results['langauge']=='EL')]
    cis = {'EN': (en_ci.Coefficient.item(), en_ci.CI_lower.item(), en_ci.CI_upper.item()),
            'EL': (el_ci.Coefficient.item(), el_ci.CI_lower.item(), el_ci.CI_upper.item()),}
    cis = pd.DataFrame(cis).T
    cis.columns = ['coef', 'lower', 'upper']
    cis.insert(0, 'language', cis.index)

    axes[i*2+1].errorbar(cis['language'], cis['coef'], 
                        yerr=[cis['coef'] - cis['lower'], cis['upper'] - cis['coef']], 
                        fmt='o', capsize=5, ms=3, elinewidth=2, markeredgewidth=2,
                        ecolor='#e18727ff', mfc='#7e6148ff', mec='#7e6148ff')

    axes[i*2+1].set_xlim(-0.5, 1.5)
    axes[i*2+1].set_title(f'({chr(65+i*2+1)}) Error bar for coefficients for {feat}', fontsize=20)
    axes[i*2+1].set_xticklabels(axes[i*2+1].get_xticklabels(), fontsize=18)
    axes[i*2+1].set_yticklabels(axes[i*2+1].get_yticklabels(), fontsize=18)
    axes[i*2+1].axhline(en_ci.Coefficient.item(), color='gray', linestyle='dashed', linewidth=2)
    
fig.savefig('MMSE_correlation.jpg', dpi=600, bbox_inches='tight')



















